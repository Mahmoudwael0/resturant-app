#libraries
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

#to start
resturant_page = tk.Tk()

#page details
resturant_page.geometry('960x540+480+270')
resturant_page.config(background='#420c09')
resturant_page.minsize(960,540)
resturant_page.title('Los pollos hermanos x Pizza hut')
resturant_page.iconbitmap('img\lph x ph.ico')

#page logo
page_logo_path='img/lph x ph.png'
page_logo_img=Image.open(page_logo_path)
page_resized=page_logo_img.resize((400, 210))
page_logo=ImageTk.PhotoImage(page_resized)
pl=tk.Canvas(resturant_page, width=960, height=540, bg="#420c09", highlightthickness=0)
pl.pack(fill="both", expand=True)
pl.create_image(100, -50, image=page_logo, anchor="nw")

#back button
back_button=tk.Button(resturant_page, text="⇐", font=("Arial", 15), width=3, height=3, bg="#420c09", fg="white", command=resturant_page.destroy)
back_button.place(x=0, y=0)




#hawaiian pizza
#frame photo
hawaiian_path='img\Hawaiian pizza.png'
hawaiian_img=Image.open(hawaiian_path)
hawaiian_resized=hawaiian_img.resize((200,300))
hawaiian_photo=ImageTk.PhotoImage(hawaiian_resized)
pl.create_image(550, 300, image=hawaiian_photo, anchor="nw")

#backend ha1
counter2=0
def update_labelha():
    label4.config(text=str(counter2))
def incrementha():
    global counter2
    counter2 += 1
    update_labelha()
def decrementha():
    global counter2
    if counter2 > 0:  
        counter2 -= 1
    update_labelha()

#backend ha2
def save_to_excel2(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label4=tk.Label(resturant_page, text=str(counter2), font=("Arial", 8),width=1, height=1)
label4.place(x=700,y=518)

#up button
button_up3=tk.Button(resturant_page, text="▲", command=incrementha, font=("Arial", 4),width=1, height=1)
button_up3.place(x=715,y=518)

#down button
button_down3=tk.Button(resturant_page, text="▼", command=decrementha, font=("Arial", 4),width=1, height=1)
button_down3.place(x=715,y=529)

#price
label5=tk.Label(resturant_page, text="140EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label5.place(x=590,y=511)

#product name
hawaiian=tk.Label(resturant_page, text="Hawaiian", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
hawaiian.place(x=560,y=470)

#buy button
button3 = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel2(counter, "Hawaiian", 140), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button3.place(x=575, y=550)

#vegetarian supreme pizza
#frame photo
vegetarian_path='img\Vegetarian supreme pizza.png'
vegetarian_img=Image.open(vegetarian_path)
vegetarian_resized=vegetarian_img.resize((200,300))
vegetarian_photo=ImageTk.PhotoImage(vegetarian_resized)
pl.create_image(800, 300, image=vegetarian_photo, anchor="nw")

#backend ve1
counter3=0
def update_labelve():
    label6.config(text=str(counter3))
def incrementve():
    global counter3
    counter3 += 1
    update_labelve()
def decrementve():
    global counter3
    if counter3 > 0:  
        counter3 -= 1
    update_labelve()
    
#backend ve2
def save_to_excel3(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label6=tk.Label(resturant_page, text=str(counter3), font=("Arial", 8),width=1, height=1)
label6.place(x=950,y=518)

#up button
button_up4=tk.Button(resturant_page, text="▲", command=incrementve, font=("Arial", 4),width=1, height=1)
button_up4.place(x=965,y=518)

#down button
button_down4=tk.Button(resturant_page, text="▼", command=decrementve, font=("Arial", 4),width=1, height=1)
button_down4.place(x=965,y=529)

#price
label7=tk.Label(resturant_page, text="160EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label7.place(x=840,y=511)

#product name
vegetarian=tk.Label(resturant_page, text="Vegetarian", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
vegetarian.place(x=810,y=470)

#buy button
button4 = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel3(counter, "Vegetarian", 160), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button4.place(x=825, y=550)

#chicken supreme pizza
#frame photo
chicken_path='img\Chicken supreme pizza.png'
chicken_img=Image.open(chicken_path)
chicken_resized=chicken_img.resize((200,300))
chicken_photo=ImageTk.PhotoImage(chicken_resized)
pl.create_image(1050, 300, image=chicken_photo, anchor="nw")

#backend ch1
counter4=0
def update_labelch():
    label8.config(text=str(counter4))
def incrementch():
    global counter4
    counter4 += 1
    update_labelch()
def decrementch():
    global counter4
    if counter4 > 0:  
        counter4 -= 1
    update_labelch()

#backend ch2
def save_to_excel4(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label8=tk.Label(resturant_page, text=str(counter4), font=("Arial", 8),width=1, height=1)
label8.place(x=1200,y=518)

#up button
button_up5=tk.Button(resturant_page, text="▲", command=incrementch, font=("Arial", 4),width=1, height=1)
button_up5.place(x=1215,y=517)

#down button
button_down5=tk.Button(resturant_page, text="▼", command=decrementch, font=("Arial", 4),width=1, height=1)
button_down5.place(x=1215,y=528)

#price
label9=tk.Label(resturant_page, text="180EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label9.place(x=1100,y=518)

#product name
chicken=tk.Label(resturant_page, text="Chicken", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
chicken.place(x=1060,y=470)

#buy button
button5 = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel4(counter, "Chicken", 180), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button5.place(x=1075, y=550)

#beef supreme pizza
#frame photo
beef_path='img\Beef supreme pizza.png'
beef_img=Image.open(beef_path)
beef_resized=beef_img.resize((200,300))
beef_photo=ImageTk.PhotoImage(beef_resized)
pl.create_image(1300, 300, image=beef_photo, anchor="nw")

#backend be1
counter5=0
def update_labelbe():
    label10.config(text=str(counter5))
def incrementbe():
    global counter5
    counter5 += 1
    update_labelbe()
def decrementbe():
    global counter5
    if counter5 > 0:  
        counter5 -= 1
    update_labelbe()

#backend be2
def save_to_excel5(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label10=tk.Label(resturant_page, text=str(counter5), font=("Arial", 8),width=1, height=1)
label10.place(x=1445,y=518)

#up button
button_up6=tk.Button(resturant_page, text="▲", command=incrementbe, font=("Arial", 4),width=1, height=1)
button_up6.place(x=1460,y=518)

#down button
button_down6=tk.Button(resturant_page, text="▼", command=decrementbe, font=("Arial", 4),width=1, height=1)
button_down6.place(x=1460,y=529)

#price
label11=tk.Label(resturant_page, text="200EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label11.place(x=1340,y=511)

#product name
beef=tk.Label(resturant_page, text="Beef", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
beef.place(x=1310,y=470)

#buy button
button6 = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel5(counter, "Beef", 200), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button6.place(x=1330, y=550)

#margherita pizza
#frame photo
margherita_path='img\Margherita pizza.png'
margherita_img=Image.open(margherita_path)
margherita_resized=margherita_img.resize((200,300))
margherita_photo=ImageTk.PhotoImage(margherita_resized)
pl.create_image(55, 300, image=margherita_photo, anchor="nw")

#backend ma1
counter6=0
def update_labelma():
    label.config(text=str(counter6))
def incrementma():
    global counter6
    counter6 += 1
    update_labelma()
def decrementma():
    global counter6
    if counter6 > 0:  
        counter6 -= 1
    update_labelma()
    
#backend ma2
def save_to_excel6(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label=tk.Label(resturant_page, text=str(counter6), font=("Arial", 8),width=1, height=1)
label.place(x=200,y=518)
#up button
button_up=tk.Button(resturant_page, text="▲", command=incrementma, font=("Arial", 4),width=1, height=1)
button_up.place(x=215,y=518)
#down button
button_down=tk.Button(resturant_page, text="▼", command=decrementma, font=("Arial", 4),width=1, height=1)
button_down.place(x=215,y=529)
#price
label1=tk.Label(resturant_page, text="99EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label1.place(x=90,y=511)
#product name
margherita=tk.Label(resturant_page, text="Margherita", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
margherita.place(x=60,y=470)

#buy button
button = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel6(counter6, "Margherita", 99), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button.place(x=69, y=550)

#pepperoni pizza
#frame photo
pepperoni_path='img\Pepperoni pizza.png'
pepperoni_img=Image.open(pepperoni_path)
pepperoni_resized=pepperoni_img.resize((200,300))
pepperoni_photo=ImageTk.PhotoImage(pepperoni_resized)
pl.create_image(300, 300, image=pepperoni_photo, anchor="nw")

#backend pe1
counter=0
def update_labelpe():
    label2.config(text=str(counter))
def incrementpe():
    global counter
    counter += 1
    update_labelpe()
def decrementpe():
    global counter
    if counter > 0:  
        counter -= 1
    update_labelpe()

#backend pe2
def save_to_excel1(quantity, product_name, price):
    try:
        workbook = load_workbook("lph x phsh.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Quantity"
        sheet["B1"] = "Product Name"
        sheet["C1"] = "Price"

    row = sheet.max_row + 1
    sheet[f"A{row}"] = quantity
    sheet[f"B{row}"] = product_name
    sheet[f"C{row}"] = price
    workbook.save("lph x phsh.xlsx")

#counter
label2=tk.Label(resturant_page, text=str(counter), font=("Arial", 8),width=1, height=1)
label2.place(x=450,y=518)

#up button
button_up2=tk.Button(resturant_page, text="▲", command=incrementpe, font=("Arial", 4),width=1, height=1)
button_up2.place(x=465,y=518)

#down button
button_down2=tk.Button(resturant_page, text="▼", command=decrementpe, font=("Arial", 4),width=1, height=1)
button_down2.place(x=465,y=529)

#price
label3=tk.Label(resturant_page, text="120EGP", font=("Arial", 10),width=7, height=2, bg="#420c09", fg="white")
label3.place(x=340,y=511)

#product name
pepperoni=tk.Label(resturant_page, text="Pepperoni", font=("Arial", 12),width=20, height=2, bg="#420c09", fg="white")
pepperoni.place(x=310,y=470)

#buy button
button2 = tk.Button(resturant_page, text="buy", command=lambda: save_to_excel1(counter, "Pepperoni", 120), font=("Arial", 10), width=10, height=1, bg="green", fg="white")
button2.place(x=325, y=550)

resturant_page.mainloop()